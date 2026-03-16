"""
Equity Beta Calculator — Global Edition
Professional light theme · Inter font · Auto index routing
Run: streamlit run beta_app.py
"""

import warnings
warnings.filterwarnings("ignore")

import streamlit as st
from streamlit_searchbox import st_searchbox
import pandas as pd
import numpy as np
import yfinance as yf
import requests
from datetime import date, timedelta, datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(
    page_title="Beta Calculator",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════════════
#  GLOBAL INDEX ROUTING
#  Yahoo Finance suffix → (index_ticker, index_name, country/region)
# ══════════════════════════════════════════════════════════════════════════════
SUFFIX_INDEX = {
    # India
    ".NS":  ("^NSEI",      "Nifty 50",          "India (NSE)"),
    ".BO":  ("^BSESN",     "Sensex",             "India (BSE)"),
    # United States
    "":     ("^GSPC",      "S&P 500",            "United States"),
    ".O":   ("^GSPC",      "S&P 500",            "United States"),
    ".Q":   ("^GSPC",      "S&P 500",            "United States"),
    # United Kingdom
    ".L":   ("^FTSE",      "FTSE 100",           "United Kingdom"),
    # Germany
    ".DE":  ("^GDAXI",     "DAX 40",             "Germany"),
    ".F":   ("^GDAXI",     "DAX 40",             "Germany"),
    # France
    ".PA":  ("^FCHI",      "CAC 40",             "France"),
    # Japan
    ".T":   ("^N225",      "Nikkei 225",         "Japan"),
    # Hong Kong
    ".HK":  ("^HSI",       "Hang Seng",          "Hong Kong"),
    # China
    ".SS":  ("000001.SS",  "SSE Composite",      "China (Shanghai)"),
    ".SZ":  ("399001.SZ",  "SZSE Component",     "China (Shenzhen)"),
    # Australia
    ".AX":  ("^AXJO",      "ASX 200",            "Australia"),
    # Canada
    ".TO":  ("^GSPTSE",    "S&P/TSX Composite",  "Canada"),
    ".V":   ("^GSPTSE",    "S&P/TSX Composite",  "Canada"),
    # Singapore
    ".SI":  ("^STI",       "Straits Times Index","Singapore"),
    # South Korea
    ".KS":  ("^KS11",      "KOSPI",              "South Korea"),
    ".KQ":  ("^KQ11",      "KOSDAQ",             "South Korea"),
    # Brazil
    ".SA":  ("^BVSP",      "Bovespa",            "Brazil"),
    # Italy
    ".MI":  ("FTSEMIB.MI", "FTSE MIB",           "Italy"),
    # Spain
    ".MC":  ("^IBEX",      "IBEX 35",            "Spain"),
    # Netherlands
    ".AS":  ("^AEX",       "AEX",                "Netherlands"),
    # Switzerland
    ".SW":  ("^SSMI",      "SMI",                "Switzerland"),
    # Sweden
    ".ST":  ("^OMX",       "OMX Stockholm 30",   "Sweden"),
    # Norway
    ".OL":  ("^OSEAX",     "Oslo Bors All Share","Norway"),
    # Denmark
    ".CO":  ("^OMXC25",    "OMX Copenhagen 25",  "Denmark"),
    # Finland
    ".HE":  ("^OMXH25",    "OMX Helsinki 25",    "Finland"),
    # New Zealand
    ".NZ":  ("^NZ50",      "NZX 50",             "New Zealand"),
    # Portugal
    ".LS":  ("^PSI20",     "PSI 20",             "Portugal"),
    # Belgium
    ".BR":  ("^BFX",       "BEL 20",             "Belgium"),
    # Austria
    ".VI":  ("^ATX",       "ATX",                "Austria"),
    # Mexico
    ".MX":  ("^MXX",       "IPC Mexico",         "Mexico"),
    # South Africa
    ".JO":  ("^J203.JO",   "JSE All Share",      "South Africa"),
    # Saudi Arabia
    ".SR":  ("^TASI.SR",   "Tadawul All Share",  "Saudi Arabia"),
    # UAE
    ".AD":  ("^FTFADGI",   "ADX General",        "UAE"),
    # Taiwan
    ".TW":  ("^TWII",      "TAIEX",              "Taiwan"),
    ".TWO": ("^TWOII",     "TPEX",               "Taiwan"),
    # Thailand
    ".BK":  ("^SET.BK",    "SET Index",          "Thailand"),
    # Malaysia
    ".KL":  ("^KLSE",      "FTSE Bursa Malaysia","Malaysia"),
    # Indonesia
    ".JK":  ("^JKSE",      "IDX Composite",      "Indonesia"),
    # Philippines
    ".PS":  ("PSEi.PS",    "PSEi",               "Philippines"),
    # Israel
    ".TA":  ("^TA125.TA",  "TA-125",             "Israel"),
}

def get_index_for_symbol(yf_symbol: str):
    """Return (index_ticker, index_name, region) for any Yahoo Finance symbol."""
    sym = yf_symbol.upper()
    # Try longest suffix match first
    for suffix in sorted(SUFFIX_INDEX.keys(), key=len, reverse=True):
        if suffix and sym.endswith(suffix.upper()):
            return SUFFIX_INDEX[suffix]
    # No suffix = US stock
    return SUFFIX_INDEX[""]


# ══════════════════════════════════════════════════════════════════════════════
#  CSS — Professional light theme, Inter font
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    background-color: #F8F9FA !important;
    color: #1A1A2E !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #FFFFFF !important;
    border-right: 1px solid #E5E7EB !important;
}
[data-testid="stSidebar"] * { color: #374151 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stDateInput label { font-size: 0.78rem !important; font-weight: 500 !important; color: #6B7280 !important; }

/* ── Top header bar ── */
.top-bar {
    background: #FFFFFF;
    border-bottom: 1px solid #E5E7EB;
    padding: 20px 0 16px 0;
    margin-bottom: 28px;
}
.app-title {
    font-size: 1.5rem;
    font-weight: 700;
    color: #111827;
    letter-spacing: -0.3px;
}
.app-subtitle {
    font-size: 0.82rem;
    color: #6B7280;
    font-weight: 400;
    margin-top: 2px;
}

/* ── Section labels ── */
.section-label {
    font-size: 0.7rem;
    font-weight: 600;
    color: #6B7280;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 10px;
}

/* ── Cards ── */
.card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 8px;
}
.card-selected {
    background: #F0FDF4;
    border: 1px solid #BBF7D0;
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 6px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}

/* ── Beta result cards ── */
.beta-card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 12px;
    padding: 22px 18px;
    text-align: center;
    transition: box-shadow 0.15s ease;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.beta-card:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.08); }
.beta-ticker   { font-size: 0.7rem; font-weight: 600; color: #9CA3AF; letter-spacing: 0.06em; text-transform: uppercase; margin-bottom: 6px; }
.beta-value    { font-size: 2.2rem; font-weight: 700; letter-spacing: -1px; line-height: 1; }
.beta-category { font-size: 0.72rem; font-weight: 500; color: #6B7280; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 6px; }
.beta-r2       { font-size: 0.75rem; color: #9CA3AF; margin-top: 5px; }
.beta-region   { font-size: 0.68rem; color: #6B7280; margin-top: 5px; }

/* ── Tags ── */
.tag {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.7rem;
    font-weight: 600;
    margin-right: 4px;
}
.tag-india  { background: #EFF6FF; color: #1D4ED8; }
.tag-us     { background: #F0FDF4; color: #166534; }
.tag-uk     { background: #FFF7ED; color: #9A3412; }
.tag-eu     { background: #FAF5FF; color: #6B21A8; }
.tag-asia   { background: #FFF1F2; color: #9F1239; }
.tag-other  { background: #F9FAFB; color: #374151; }

/* ── Hint box ── */
.hint {
    background: #EFF6FF;
    border: 1px solid #BFDBFE;
    border-radius: 8px;
    padding: 10px 14px;
    font-size: 0.82rem;
    color: #1D4ED8;
    margin-bottom: 14px;
}

/* ── Status boxes ── */
.status-ok  { background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 8px; padding: 10px 14px; color: #166534; font-size: 0.82rem; font-weight: 500; }
.status-err { background: #FFF1F2; border: 1px solid #FECDD3; border-radius: 8px; padding: 10px 14px; color: #9F1239; font-size: 0.82rem; font-weight: 500; }

/* ── Divider ── */
.divider { height: 1px; background: #E5E7EB; margin: 22px 0; }

/* ── Table ── */
.stDataFrame { border: 1px solid #E5E7EB !important; border-radius: 10px !important; overflow: hidden; }

/* ── Buttons ── */
.stButton > button {
    background: #1D4ED8 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.88rem !important;
    padding: 10px 20px !important;
    transition: background 0.15s !important;
    width: 100%;
}
.stButton > button:hover { background: #1E40AF !important; }
.stButton > button:disabled { background: #D1D5DB !important; color: #9CA3AF !important; }

/* Clear button */
.clear-btn > button {
    background: #FFFFFF !important;
    color: #6B7280 !important;
    border: 1px solid #D1D5DB !important;
    border-radius: 8px !important;
}
.clear-btn > button:hover { background: #F9FAFB !important; border-color: #9CA3AF !important; }

/* ── Download button ── */
.stDownloadButton > button {
    background: #FFFFFF !important;
    color: #1D4ED8 !important;
    border: 1px solid #BFDBFE !important;
    border-radius: 8px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    width: 100%;
}
.stDownloadButton > button:hover { background: #EFF6FF !important; }

/* ── Progress ── */
.stProgress > div > div > div { background: #1D4ED8 !important; }

/* ── Inputs ── */
div[data-testid="stTextInput"] input,
div[data-testid="stSelectbox"] select {
    background: #FFFFFF !important;
    border: 1px solid #D1D5DB !important;
    border-radius: 8px !important;
    color: #111827 !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.9rem !important;
}
div[data-testid="stTextInput"] input:focus {
    border-color: #1D4ED8 !important;
    box-shadow: 0 0 0 3px rgba(29,78,216,0.1) !important;
}

/* ── Searchbox ── */
div[data-testid="stSearchbox"] > div > div,
div[data-testid="stSearchbox"] > div,
div[data-testid="stSearchbox"] {
    border: 1px solid #D1D5DB !important;
    border-radius: 8px !important;
    background: #FFFFFF !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
}
div[data-testid="stSearchbox"] input,
div[data-testid="stSearchbox"] input::placeholder {
    font-family: 'Inter', sans-serif !important;
    font-size: 0.92rem !important;
    color: #111827 !important;
    background: #FFFFFF !important;
    -webkit-text-fill-color: #111827 !important;
}
div[data-testid="stSearchbox"] input:focus {
    border-color: #1D4ED8 !important;
    box-shadow: 0 0 0 3px rgba(29,78,216,0.1) !important;
    background: #FFFFFF !important;
}
/* Force all inner divs of searchbox to white */
div[data-testid="stSearchbox"] * {
    background-color: #FFFFFF !important;
    color: #111827 !important;
}
div[data-testid="stSearchbox"] ul,
div[data-testid="stSearchbox"] ul * {
    background-color: #FFFFFF !important;
}
div[data-testid="stSearchbox"] li:hover,
div[data-testid="stSearchbox"] li:hover * {
    background-color: #EFF6FF !important;
    color: #1D4ED8 !important;
}
div[data-testid="stSearchbox"] ul {
    background: #FFFFFF !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 8px !important;
    box-shadow: 0 8px 24px rgba(0,0,0,0.12) !important;
    margin-top: 4px !important;
}
div[data-testid="stSearchbox"] li {
    font-family: 'Inter', sans-serif !important;
    font-size: 0.86rem !important;
    color: #374151 !important;
    padding: 10px 14px !important;
    border-bottom: 1px solid #F3F4F6 !important;
}
div[data-testid="stSearchbox"] li:hover { background: #EFF6FF !important; color: #1D4ED8 !important; }

/* ── Sidebar inputs — force light ── */
[data-testid="stSidebar"] div[data-testid="stDateInput"] input,
[data-testid="stSidebar"] input {
    background: #F9FAFB !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 6px !important;
    color: #111827 !important;
    font-size: 0.85rem !important;
}
[data-testid="stSidebar"] div[data-testid="stSelectbox"] > div,
[data-testid="stSidebar"] div[data-baseweb="select"] > div,
[data-testid="stSidebar"] div[data-baseweb="select"] {
    background: #F9FAFB !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 6px !important;
    color: #111827 !important;
}
[data-testid="stSidebar"] div[data-baseweb="select"] span,
[data-testid="stSidebar"] div[data-baseweb="select"] div {
    color: #111827 !important;
    background: #F9FAFB !important;
}
/* Sidebar dropdown menu */
[data-testid="stSidebar"] ul[data-baseweb="menu"],
[data-testid="stSidebar"] ul {
    background: #FFFFFF !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] li {
    color: #111827 !important;
    background: #FFFFFF !important;
}
[data-testid="stSidebar"] li:hover {
    background: #EFF6FF !important;
    color: #1D4ED8 !important;
}
/* Sidebar button — white with blue border */
[data-testid="stSidebar"] .stButton > button {
    background: #FFFFFF !important;
    color: #1D4ED8 !important;
    border: 1px solid #BFDBFE !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    font-size: 0.88rem !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #EFF6FF !important;
    border-color: #1D4ED8 !important;
}
/* Sidebar warning */
[data-testid="stSidebar"] .stWarning {
    background: #FFFBEB !important;
    color: #92400E !important;
    border-radius: 6px !important;
}

/* ── Warning ── */
.stWarning { border-radius: 8px !important; }

#MainMenu { visibility: hidden; }
footer     { visibility: hidden; }
header     { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  CLOUD-COMPATIBLE SESSION (curl_cffi for Streamlit Cloud / AWS)
# ══════════════════════════════════════════════════════════════════════════════
try:
    from curl_cffi import requests as _curl
    _SESSION = _curl.Session(impersonate="chrome110")
    _HAS_CURL = True
except Exception:
    _SESSION = None
    _HAS_CURL = False

def _get_ticker(symbol: str):
    if _HAS_CURL and _SESSION is not None:
        return yf.Ticker(symbol, session=_SESSION)
    return yf.Ticker(symbol)


# ══════════════════════════════════════════════════════════════════════════════
#  DATA FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=120, show_spinner=False)
def live_search(query: str) -> list:
    if not query or len(query.strip()) < 1:
        return []
    try:
        if _HAS_CURL and _SESSION is not None:
            results = yf.Search(query.strip(), max_results=12, news_count=0, session=_SESSION)
        else:
            results = yf.Search(query.strip(), max_results=12, news_count=0)
        quotes = results.quotes if hasattr(results, "quotes") else []
        found  = []
        for q in quotes:
            sym  = q.get("symbol", "")
            # Accept any equity globally — not just .NS/.BO
            qtype = q.get("quoteType", "").upper()
            if qtype not in ("EQUITY", "ETF", ""):
                continue
            name = q.get("longname") or q.get("shortname") or sym
            idx_ticker, idx_name, region = get_index_for_symbol(sym)
            found.append({
                "name":       name,
                "symbol":     sym,
                "index_yf":   idx_ticker,
                "index_name": idx_name,
                "region":     region,
            })
        return found
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_prices(yf_symbol: str, start: str, end: str) -> pd.Series:
    try:
        tkr = _get_ticker(yf_symbol)
        raw = tkr.history(start=start, end=end, auto_adjust=True)
        if raw.empty:
            return pd.Series(dtype=float)
        close = raw["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        close.index = pd.to_datetime(close.index).normalize()
        return close.astype(float)
    except Exception:
        return pd.Series(dtype=float)


def calc_beta(comp_ret, idx_ret):
    df = pd.concat([comp_ret, idx_ret], axis=1).dropna()
    df.columns = ["c", "i"]
    n = len(df)
    if n < 20:
        return None, None, None, None, None, n
    x, y     = df["i"].values, df["c"].values
    slope, _ = np.polyfit(x, y, 1)
    corr     = float(np.corrcoef(x, y)[0, 1])
    # Annualised volatility: daily StdDev × sqrt(252)
    stock_vol = float(df["c"].std() * np.sqrt(252))
    index_vol = float(df["i"].std() * np.sqrt(252))
    return round(float(slope), 4), round(corr**2, 4), round(corr, 4), round(stock_vol, 4), round(index_vol, 4), n


def beta_style(b):
    if b is None:  return "#9CA3AF", "N/A",          "#F9FAFB"
    if b < 0:      return "#DC2626", "Negative",     "#FFF1F2"
    if b < 0.8:    return "#2563EB", "Defensive",    "#EFF6FF"
    if b <= 1.2:   return "#059669", "Market-like",  "#F0FDF4"
    if b <= 1.5:   return "#D97706", "Moderate",     "#FFFBEB"
    return             "#DC2626",    "Aggressive",   "#FFF1F2"


def region_tag(region: str) -> str:
    r = region.lower()
    if "india"    in r: return "tag-india"
    if "united s" in r: return "tag-us"
    if "united k" in r: return "tag-uk"
    if any(x in r for x in ["germany","france","italy","spain","nether","sweden","norway","denmark","finland","belgium","austria","portugal"]): return "tag-eu"
    if any(x in r for x in ["japan","china","hong kong","korea","singapore","taiwan","thailand","malaysia","indonesia","philippines"]): return "tag-asia"
    return "tag-other"


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
_CLR = {"h_bg":"1D4ED8","s_bg":"3B82F6","alt":"F8FAFF","wht":"FFFFFF",
        "grn":"DCFCE7","amb":"FEF3C7","bdr":"E5E7EB"}
_T   = Side(style="thin", color=_CLR["bdr"])
_BDR = Border(left=_T, right=_T, top=_T, bottom=_T)

def _xhdr(cell, bg=None, sz=10):
    cell.font      = Font(name="Calibri", size=sz, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", start_color=bg or _CLR["h_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BDR

def _xcell(cell, value, bg="FFFFFF", fmt=None, bold=False, right=False):
    cell.value     = value
    cell.font      = Font(name="Calibri", size=10 if bold else 9, bold=bold)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.border    = _BDR
    cell.alignment = Alignment(horizontal="right" if right else "center", vertical="center")
    if fmt:
        cell.number_format = fmt

def beta_style_excel(b):
    """Return (color_hex, category_label, fill_hex) for Excel use."""
    if b is None:  return "9CA3AF", "N/A",          "FFFFFF"
    if b < 0:      return "DC2626", "Negative",     "FFF1F2"
    if b < 0.8:    return "2563EB", "Defensive",    "EFF6FF"
    if b <= 1.2:   return "059669", "Market-like",  "F0FDF4"
    if b <= 1.5:   return "D97706", "Moderate",     "FFFBEB"
    return             "DC2626",    "Aggressive",   "FFF1F2"

def build_excel(results: list, start_str: str, end_str: str) -> bytes:
    wb   = Workbook()
    s_ws = wb.active; s_ws.title = "Summary"
    m_ws = wb.create_sheet("Methodology")

    s_ws.sheet_view.showGridLines = False
    s_ws.row_dimensions[1].height = 36
    s_ws.row_dimensions[3].height = 26

    s_ws.merge_cells("A1:M1")
    c = s_ws["A1"]
    c.value = "EQUITY BETA ANALYSIS — GLOBAL"
    c.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=_CLR["h_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    s_ws.merge_cells("A2:M2")
    c = s_ws["A2"]
    c.value = "Period: " + start_str + " to " + end_str + "  |  Run: " + datetime.today().strftime("%d-%b-%Y") + "  |  Beta = SLOPE(Stock Returns, Index Returns)"
    c.font  = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
    c.fill  = PatternFill("solid", start_color=_CLR["s_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["S.No.","Company","Ticker","Region","Benchmark Index","Index Ticker",
            "Start Date","End Date","Obs.",
            "Beta","Beta Category",
            "Stock Vol (Ann.)","Index Vol (Ann.)",
            "R²","Correlation","Remarks"]
    for j, h in enumerate(hdrs):
        _xhdr(s_ws.cell(row=3, column=j+1, value=h))

    for i, r in enumerate(results):
        row = 4+i; bg = _CLR["alt"] if i%2 else _CLR["wht"]
        _, beta_cat, _ = beta_style_excel(r.get("beta"))
        sv = r.get("stock_vol")
        iv = r.get("index_vol")
        vals = [i+1, r["name"], r["symbol"], r["region"], r["index_name"], r["index_yf"],
                r.get("start_date","—"), r.get("end_date","—"), r.get("n_obs",0),
                r.get("beta"), beta_cat,
                sv, iv,
                r.get("r2"), r.get("corr"), r.get("error") or ""]
        for j, v in enumerate(vals):
            _xcell(s_ws.cell(row=row, column=j+1), v, bg=bg)

        # Beta — colour coded
        bc = s_ws.cell(row=row, column=10)
        bc.number_format = "0.00"; bc.font = Font(name="Calibri", size=10, bold=True)
        if r.get("beta") is not None:
            b = r["beta"]
            fc = _CLR["grn"] if 0.8<=b<=1.5 else (_CLR["amb"] if (b>1.5 or b<0) else _CLR["wht"])
            bc.fill = PatternFill("solid", start_color=fc)

        # Volatility — percentage format
        vc = s_ws.cell(row=row, column=12)
        vc.number_format = "0.00%"
        vc.font = Font(name="Calibri", size=9, bold=True,
                       color="1D4ED8" if sv else "9CA3AF")
        ic = s_ws.cell(row=row, column=13)
        ic.number_format = "0.00%"

        s_ws.cell(row=row, column=9).number_format  = "#,##0"
        s_ws.cell(row=row, column=14).number_format = "0.0000"
        s_ws.cell(row=row, column=15).number_format = "0.0000"

    note = 5+len(results)
    s_ws.merge_cells("A"+str(note)+":P"+str(note))
    s_ws["A"+str(note)].value = ("Green = Beta 0.80–1.50  |  Amber = >1.50 or Negative  |  "
                                  "Volatility = Annualised StdDev of Daily Returns × √252  |  "
                                  "Beta = SLOPE(Stock Returns, Index Returns)")
    s_ws["A"+str(note)].font = Font(name="Calibri", size=8, italic=True)
    s_ws["A"+str(note)].alignment = Alignment(horizontal="left")

    for col, w in [("A",5),("B",26),("C",12),("D",16),("E",20),("F",10),
                   ("G",13),("H",13),("I",8),("J",8),("K",14),
                   ("L",14),("M",14),("N",8),("O",12),("P",20)]:
        s_ws.column_dimensions[col].width = w
    s_ws.freeze_panes = "A4"

    for r in results:
        if r.get("aligned") is None:
            continue
        aligned   = r["aligned"]
        idx_label = r["index_name"]
        dws = wb.create_sheet(r["symbol"][:28])
        dws.sheet_view.showGridLines = False
        last_r = 3+len(aligned)

        dws.merge_cells("A1:I1")
        c = dws["A1"]
        c.value = r["name"] + "  vs  " + idx_label + "  (" + r["region"] + ")"
        c.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color=_CLR["h_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        dws.row_dimensions[1].height = 26

        dws.merge_cells("A2:D2")
        c = dws["A2"]
        c.value = "Beta = SLOPE(I4:I"+str(last_r)+", H4:H"+str(last_r)+")  —  verify cell E2"
        c.font  = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color=_CLR["s_bg"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        dws.merge_cells("E2:I2")
        c = dws["E2"]
        c.value = "=SLOPE(I4:I"+str(last_r)+",H4:H"+str(last_r)+")"
        c.font  = Font(name="Calibri", size=11, bold=True)
        c.fill  = PatternFill("solid", start_color=_CLR["grn"])
        c.number_format = "0.0000"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BDR

        h2 = ["Date", idx_label+" Price", idx_label+" Return", "",
               r["name"]+" Price", r["name"]+" Return", "",
               "Index Return (x)", "Stock Return (y)"]
        for j, h in enumerate(h2):
            _xhdr(dws.cell(row=3, column=j+1, value=h), bg=_CLR["s_bg"])

        fmt_map = {1:"#,##0.00",2:"0.00%",4:"#,##0.00",5:"0.00%",7:"0.00%",8:"0.00%"}
        for ii, (dt, row_d) in enumerate(aligned.iterrows()):
            rr = 4+ii; bg = _CLR["alt"] if ii%2 else _CLR["wht"]
            rv = [dt.strftime("%d-%b-%Y"),
                  row_d["index_price"], row_d["index_return"], "",
                  row_d["comp_price"],  row_d["comp_return"],  "",
                  row_d["index_return"],row_d["comp_return"]]
            for j, v in enumerate(rv):
                _xcell(dws.cell(row=rr, column=j+1), v, bg=bg,
                       fmt=fmt_map.get(j), right=(j>0 and j not in (3,6)))

        for col, w in [("A",14),("B",18),("C",18),("D",4),
                       ("E",18),("F",20),("G",4),("H",22),("I",22)]:
            dws.column_dimensions[col].width = w
        dws.freeze_panes = "A4"

    m_ws.sheet_view.showGridLines = False
    m_ws.column_dimensions["A"].width = 100
    mrows = [
        ("BETA CALCULATION — METHODOLOGY (GLOBAL)",          True,  _CLR["h_bg"], "FFFFFF", 13),
        ("",                                                  False, None,         None,     10),
        ("1.  DATA SOURCE & GLOBAL INDEX ROUTING",           True,  _CLR["s_bg"], "FFFFFF", 11),
        ("Data via Yahoo Finance API. Each stock is automatically matched to its country's broad market index.", False, None, None, 10),
        ("Examples: US stocks → S&P 500  |  UK → FTSE 100  |  Germany → DAX  |  Japan → Nikkei 225", False, None, None, 10),
        ("India NSE (.NS) → Nifty 50  |  India BSE (.BO) → Sensex  |  Hong Kong (.HK) → Hang Seng", False, None, None, 10),
        ("Prices: Adjusted Closing — auto-adjusted for splits and dividends.",     False, None, None, 10),
        ("",                                                  False, None,         None,     10),
        ("2.  RETURN CALCULATION",                           True,  _CLR["s_bg"], "FFFFFF", 11),
        ("Daily Return(t) = [Close(t) - Close(t-1)] / Close(t-1)  — identical to Excel =(B2-B1)/B1.", False, None, None, 10),
        ("Date alignment: inner join — only days where BOTH the stock and its index have data.", False, None, None, 10),
        ("",                                                  False, None,         None,     10),
        ("3.  BETA FORMULA",                                 True,  _CLR["s_bg"], "FFFFFF", 11),
        ("Beta = SLOPE(Stock Daily Returns, Benchmark Index Daily Returns)", False, None, None, 10),
        ("Python: numpy.polyfit(x, y, 1)[0]  — mathematically identical to Excel SLOPE.", False, None, None, 10),
        ("Cell E2 on each sheet contains a live =SLOPE() formula for independent audit.", False, None, None, 10),
        ("",                                                  False, None,         None,     10),
        ("4.  CAPM CAVEATS",                                 True,  _CLR["s_bg"], "FFFFFF", 11),
        ("This is LEVERED (equity) beta. Unlever: bu = bL / [1 + (1-t)(D/E)]  (Hamada equation).", False, None, None, 10),
        ("For cross-border comparables, ensure all betas use a common base currency return series.", False, None, None, 10),
        ("Currency effects are NOT stripped — returns reflect local currency performance.", False, None, None, 10),
        ("Daily returns produce higher betas than weekly/monthly due to microstructure noise.", False, None, None, 10),
    ]
    for i, (txt, bold, bg, fg, sz) in enumerate(mrows):
        ri = i+1; c = m_ws.cell(row=ri, column=1, value=txt)
        c.font = Font(name="Calibri", size=sz, bold=bold, color=fg if fg else "000000")
        if bg: c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        m_ws.row_dimensions[ri].height = 22 if bold else 17

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
if "selected" not in st.session_state:
    st.session_state.selected = []


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style='padding:20px 0 12px;'>
        <div style='font-size:1rem;font-weight:700;color:#111827;'>Beta Calculator</div>
        <div style='font-size:0.75rem;color:#6B7280;margin-top:2px;'>Global Equity · CAPM</div>
    </div>
    <div style='height:1px;background:#E5E7EB;margin-bottom:20px;'></div>
    """, unsafe_allow_html=True)

    st.markdown("<div class='section-label'>Analysis Period</div>", unsafe_allow_html=True)
    start_date = st.date_input("From", value=date(2022, 1, 1),
                               min_value=date(2005, 1, 1),
                               max_value=date.today() - timedelta(days=30),
                               label_visibility="collapsed")
    end_date   = st.date_input("To", value=date.today(),
                               min_value=date(2005, 1, 1),
                               max_value=date.today(),
                               label_visibility="collapsed")
    st.markdown(f"<div style='font-size:0.72rem;color:#9CA3AF;margin-top:4px;'>{(end_date-start_date).days} days selected</div>",
                unsafe_allow_html=True)
    if (end_date - start_date).days < 90:
        st.warning("Minimum 90 days recommended.")

    st.markdown("<div style='height:1px;background:#E5E7EB;margin:20px 0;'></div>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Quick Presets</div>", unsafe_allow_html=True)

    PRESETS = {
        "— Select —":      [],
        "S&P 500 Tech":    [("AAPL","Apple","^GSPC","S&P 500","United States"),("MSFT","Microsoft","^GSPC","S&P 500","United States"),("GOOGL","Alphabet","^GSPC","S&P 500","United States"),("AMZN","Amazon","^GSPC","S&P 500","United States"),("META","Meta","^GSPC","S&P 500","United States")],
        "Nifty IT":        [("TCS.NS","TCS","^NSEI","Nifty 50","India (NSE)"),("INFY.NS","Infosys","^NSEI","Nifty 50","India (NSE)"),("WIPRO.NS","Wipro","^NSEI","Nifty 50","India (NSE)"),("HCLTECH.NS","HCL Tech","^NSEI","Nifty 50","India (NSE)"),("TECHM.NS","Tech Mahindra","^NSEI","Nifty 50","India (NSE)")],
        "Nifty Bank":      [("HDFCBANK.NS","HDFC Bank","^NSEI","Nifty 50","India (NSE)"),("ICICIBANK.NS","ICICI Bank","^NSEI","Nifty 50","India (NSE)"),("KOTAKBANK.NS","Kotak Bank","^NSEI","Nifty 50","India (NSE)"),("AXISBANK.NS","Axis Bank","^NSEI","Nifty 50","India (NSE)"),("SBIN.NS","SBI","^NSEI","Nifty 50","India (NSE)")],
        "Global Banks":    [("JPM","JPMorgan","^GSPC","S&P 500","United States"),("HSBA.L","HSBC","^FTSE","FTSE 100","United Kingdom"),("BNP.PA","BNP Paribas","^FCHI","CAC 40","France"),("DBK.DE","Deutsche Bank","^GDAXI","DAX 40","Germany"),("8306.T","MUFG","^N225","Nikkei 225","Japan")],
        "Global Pharma":   [("JNJ","J&J","^GSPC","S&P 500","United States"),("NVS","Novartis","^SSMI","SMI","Switzerland"),("AZN.L","AstraZeneca","^FTSE","FTSE 100","United Kingdom"),("ROG.SW","Roche","^SSMI","SMI","Switzerland"),("SUNPHARMA.NS","Sun Pharma","^NSEI","Nifty 50","India (NSE)")],
        "Nifty Top 5":     [("RELIANCE.NS","Reliance","^NSEI","Nifty 50","India (NSE)"),("TCS.NS","TCS","^NSEI","Nifty 50","India (NSE)"),("HDFCBANK.NS","HDFC Bank","^NSEI","Nifty 50","India (NSE)"),("INFY.NS","Infosys","^NSEI","Nifty 50","India (NSE)"),("ICICIBANK.NS","ICICI Bank","^NSEI","Nifty 50","India (NSE)")],
    }

    preset_choice = st.selectbox("Preset", list(PRESETS.keys()), label_visibility="collapsed")
    if st.button("Load Preset", use_container_width=True):
        if preset_choice != "— Select —":
            existing = {c["symbol"] for c in st.session_state.selected}
            for sym, name, idx_yf, idx_name, region in PRESETS[preset_choice]:
                if sym not in existing:
                    st.session_state.selected.append({
                        "name":sym,"symbol":sym,"index_yf":idx_yf,
                        "index_name":idx_name,"region":region,
                    })
                    # Try to get proper name
                    for item in PRESETS[preset_choice]:
                        if item[0] == sym:
                            st.session_state.selected[-1]["name"] = item[1]
            st.rerun()

    st.markdown("<div style='height:1px;background:#E5E7EB;margin:20px 0 16px;'></div>", unsafe_allow_html=True)
    st.markdown("""
    <div style='font-size:0.72rem;color:#9CA3AF;line-height:1.9;'>
        <strong style='color:#6B7280;'>Coverage</strong><br>
        🇮🇳 NSE · BSE &nbsp;·&nbsp; 🇺🇸 NYSE · Nasdaq<br>
        🇬🇧 LSE &nbsp;·&nbsp; 🇩🇪 Xetra &nbsp;·&nbsp; 🇫🇷 Euronext<br>
        🇯🇵 TSE &nbsp;·&nbsp; 🇭🇰 HKEX &nbsp;·&nbsp; 🇨🇳 SSE<br>
        🇦🇺 ASX &nbsp;·&nbsp; 🇨🇦 TSX &nbsp;·&nbsp; 40+ markets<br><br>
        <strong style='color:#6B7280;'>Method</strong><br>
        β = SLOPE(Rᵢ, Rₘ) · Excel-identical
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

# Header
st.markdown("""
<div class='top-bar'>
    <div class='app-title'>📈 Equity Beta Calculator</div>
    <div class='app-subtitle'>Global coverage · Auto index routing · Excel-identical SLOPE methodology</div>
</div>
""", unsafe_allow_html=True)

# Two-column layout: search + selected
col_search, col_selected = st.columns([5, 4], gap="large")

with col_search:
    st.markdown("<div class='section-label'>Search Companies</div>", unsafe_allow_html=True)
    st.markdown("""
    <div class='hint'>
        Search any listed company globally — US, India, UK, Europe, Asia and more.
        The correct benchmark index is selected automatically per country.
    </div>
    """, unsafe_allow_html=True)

    def _search_fn(query: str) -> list:
        if not query or len(query.strip()) < 1:
            return []
        results = live_search(query)
        labels  = []
        for r in results:
            label = r["name"] + "  ·  " + r["symbol"] + "  ·  " + r["region"]
            labels.append(label)
        return labels

    selected_label = st_searchbox(
        _search_fn,
        placeholder="Search company name or ticker  —  e.g.  Apple  /  Reliance  /  HSBC",
        key="global_searchbox",
        clear_on_submit=True,
        debounce=200,
    )

    # Handle selection
    if selected_label:
        try:
            parts  = selected_label.split("  ·  ")
            name   = parts[0].strip()
            symbol = parts[1].strip()
            region = parts[2].strip() if len(parts) > 2 else ""
            idx_yf, idx_name, region_full = get_index_for_symbol(symbol)
            existing = {c["symbol"] for c in st.session_state.selected}
            if symbol not in existing:
                st.session_state.selected.append({
                    "name":       name,
                    "symbol":     symbol,
                    "index_yf":   idx_yf,
                    "index_name": idx_name,
                    "region":     region_full,
                })
                st.rerun()
        except Exception:
            pass

    # Index routing reference
    st.markdown("<div style='margin-top:20px;'>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Auto Index Routing</div>", unsafe_allow_html=True)
    routing_data = {
        "🇮🇳 India NSE / BSE": "Nifty 50 / Sensex",
        "🇺🇸 United States":   "S&P 500",
        "🇬🇧 United Kingdom":  "FTSE 100",
        "🇩🇪 Germany":         "DAX 40",
        "🇫🇷 France":          "CAC 40",
        "🇯🇵 Japan":           "Nikkei 225",
        "🇭🇰 Hong Kong":       "Hang Seng",
        "🇨🇳 China":           "SSE Composite",
        "🇦🇺 Australia":       "ASX 200",
        "🇨🇦 Canada":          "S&P/TSX",
        "🇸🇬 Singapore":       "STI",
        "🇧🇷 Brazil":          "Bovespa",
        "🇨🇭 Switzerland":     "SMI",
    }
    rows_html = "".join(
        f"<div style='display:flex;justify-content:space-between;padding:5px 0;"
        f"border-bottom:1px solid #F3F4F6;font-size:0.78rem;'>"
        f"<span style='color:#374151;'>{k}</span>"
        f"<span style='color:#1D4ED8;font-weight:500;'>{v}</span></div>"
        for k, v in routing_data.items()
    )
    st.markdown(
        "<div class='card' style='padding:12px 16px;'>" + rows_html + "</div>",
        unsafe_allow_html=True,
    )

with col_selected:
    count = len(st.session_state.selected)
    st.markdown(
        "<div class='section-label'>Selected Companies  " +
        (f"<span style='background:#EFF6FF;color:#1D4ED8;border-radius:10px;padding:1px 8px;font-size:0.68rem;font-weight:600;'>{count}</span>" if count else "") +
        "</div>",
        unsafe_allow_html=True,
    )

    if st.session_state.selected:
        for idx, comp in enumerate(st.session_state.selected):
            tag_cls = region_tag(comp.get("region",""))
            c1, c2  = st.columns([5, 1])
            with c1:
                st.markdown(
                    "<div class='card' style='padding:10px 14px;margin-bottom:6px;'>"
                    "<div style='font-size:0.86rem;font-weight:500;color:#111827;'>" + comp["name"] + "</div>"
                    "<div style='margin-top:4px;'>"
                    "<span class='tag " + tag_cls + "'>" + comp["symbol"] + "</span>"
                    "<span style='font-size:0.72rem;color:#9CA3AF;'>→ " + comp["index_name"] + "</span>"
                    "</div></div>",
                    unsafe_allow_html=True,
                )
            with c2:
                if st.button("✕", key="rm_" + str(idx) + "_" + comp["symbol"],
                             use_container_width=True):
                    st.session_state.selected.pop(idx)
                    st.rerun()

        with st.container():
            st.markdown("<div class='clear-btn'>", unsafe_allow_html=True)
            if st.button("Clear all", use_container_width=False):
                st.session_state.selected = []
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style='border:2px dashed #E5E7EB;border-radius:10px;padding:32px;
                    text-align:center;color:#9CA3AF;font-size:0.84rem;'>
            Search and select companies<br>to build your peer group
        </div>
        """, unsafe_allow_html=True)


# ── Calculate button ──────────────────────────────────────────────────────────
st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

btn_c, _ = st.columns([1, 3])
with btn_c:
    run = st.button(
        "Calculate Beta" + (f"  ({count} companies)" if count else ""),
        use_container_width=True,
        disabled=(count == 0),
    )


# ══════════════════════════════════════════════════════════════════════════════
#  CALCULATION
# ══════════════════════════════════════════════════════════════════════════════
if run and st.session_state.selected:
    if (end_date - start_date).days < 30:
        st.markdown("<div class='status-err'>Date range too short — minimum 30 days.</div>",
                    unsafe_allow_html=True)
        st.stop()

    start_str = start_date.strftime("%Y-%m-%d")
    end_str   = end_date.strftime("%Y-%m-%d")
    companies = st.session_state.selected

    st.markdown("<div class='section-label' style='margin-top:8px;'>Fetching Data</div>",
                unsafe_allow_html=True)
    progress = st.progress(0)
    status   = st.empty()

    # Fetch unique indices
    indices_needed = {}
    for comp in companies:
        key = comp["index_yf"]
        if key not in indices_needed:
            indices_needed[key] = {"name": comp["index_name"]}

    idx_cache = {}
    for step, (idx_yf, meta) in enumerate(indices_needed.items()):
        status.markdown(
            f"<div style='font-size:0.82rem;color:#6B7280;'>Fetching {meta['name']}...</div>",
            unsafe_allow_html=True,
        )
        px = fetch_prices(idx_yf, start_str, end_str)
        if not px.empty:
            idx_cache[idx_yf] = {"prices": px, "returns": px.pct_change().dropna(), "name": meta["name"]}
        progress.progress(int(8 * (step+1) / max(len(indices_needed), 1)))

    if not idx_cache:
        st.markdown("<div class='status-err'>Could not fetch index data. Check your internet connection.</div>",
                    unsafe_allow_html=True)
        st.stop()

    results = []
    total   = len(companies)

    for i, comp in enumerate(companies):
        idx_yf = comp["index_yf"]
        status.markdown(
            f"<div style='font-size:0.82rem;color:#6B7280;'>"
            f"Fetching {comp['symbol']}  ({i+1}/{total})...</div>",
            unsafe_allow_html=True,
        )
        prices = fetch_prices(comp["symbol"], start_str, end_str)
        progress.progress(10 + int(85*(i+1)/total))

        if prices.empty or idx_yf not in idx_cache:
            results.append({**comp, "beta":None, "r2":None, "corr":None,
                             "stock_vol":None, "index_vol":None,
                             "n_obs":0, "error":"No price data", "aligned":None})
            continue

        comp_ret = prices.pct_change().dropna()
        idx_data = idx_cache[idx_yf]
        aligned  = pd.concat({
            "index_price":  idx_data["prices"],
            "index_return": idx_data["returns"],
            "comp_price":   prices,
            "comp_return":  comp_ret,
        }, axis=1).dropna()
        aligned.columns = ["index_price","index_return","comp_price","comp_return"]

        beta, r2, corr, stock_vol, index_vol, n = calc_beta(aligned["comp_return"], aligned["index_return"])
        results.append({
            **comp,
            "beta":beta, "r2":r2, "corr":corr,
            "stock_vol":stock_vol, "index_vol":index_vol, "n_obs":n,
            "start_date": aligned.index.min().strftime("%d-%b-%Y") if len(aligned) else "—",
            "end_date":   aligned.index.max().strftime("%d-%b-%Y") if len(aligned) else "—",
            "error":      None if beta is not None else "Insufficient data",
            "aligned":    aligned if beta is not None else None,
        })

    progress.progress(100)
    ok_count = sum(1 for r in results if r["beta"] is not None)
    status.markdown(
        f"<div class='status-ok'>✓  {ok_count} of {total} companies processed successfully.</div>",
        unsafe_allow_html=True,
    )

    # ── Beta cards ────────────────────────────────────────────────────────────
    valid = [r for r in results if r["beta"] is not None]
    if valid:
        st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
        st.markdown("<div class='section-label'>Results</div>", unsafe_allow_html=True)

        n_cols = min(len(valid), 5)
        cols   = st.columns(n_cols)
        for i, r in enumerate(valid):
            color, label, bg = beta_style(r["beta"])
            tag_cls = region_tag(r.get("region",""))
            with cols[i % n_cols]:
                st.markdown(
                    f"<div class='beta-card' style='border-top:3px solid {color};'>"
                    f"<div class='beta-ticker'>{r['symbol']}</div>"
                    f"<div class='beta-value' style='color:{color};'>{round(r['beta'],2)}</div>"
                    f"<div class='beta-category'>{label}</div>"
                    f"<div class='beta-r2'>R² = {round(r['r2'],3)}</div>"
                    f"<div class='beta-r2'>Vol = {round(r['stock_vol']*100,1)}% p.a.</div>"
                    f"<div class='beta-region'>{r.get('region','')} · {r['index_name']}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    # ── Results table ─────────────────────────────────────────────────────────
    st.markdown("<div style='margin-top:20px;'>", unsafe_allow_html=True)
    rows = []
    for r in results:
        _, label, _ = beta_style(r["beta"])
        rows.append({
            "Company":      r["name"],
            "Ticker":       r["symbol"],
            "Region":       r.get("region",""),
            "Benchmark":    r["index_name"],
            "Beta":         round(r["beta"],4) if r["beta"] is not None else None,
            "Category":     label,
            "Volatility (Ann.)": f"{round(r['stock_vol']*100,2)}%" if r.get("stock_vol") else "—",
            "Index Vol (Ann.)":  f"{round(r['index_vol']*100,2)}%" if r.get("index_vol") else "—",
            "R²":           round(r["r2"],4) if r["r2"] else None,
            "Correlation":  round(r["corr"],4) if r["corr"] else None,
            "Observations": r["n_obs"],
            "Period":       r.get("start_date","—") + " → " + r.get("end_date","—"),
            "Status":       r["error"] or "✓",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True,
                 hide_index=True, height=min(420, 56+len(results)*36))

    # ── Export ────────────────────────────────────────────────────────────────
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    st.markdown("<div class='section-label'>Export</div>", unsafe_allow_html=True)

    excel_bytes = build_excel(results, start_date.strftime("%d-%b-%Y"), end_date.strftime("%d-%b-%Y"))
    fname = "Beta_Global_" + datetime.today().strftime("%Y%m%d") + ".xlsx"

    dl_col, info_col = st.columns([1, 3])
    with dl_col:
        st.download_button("⬇  Download Excel Report", data=excel_bytes,
                           file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with info_col:
        st.markdown(
            f"<div style='font-size:0.78rem;color:#6B7280;padding-top:10px;'>"
            f"{fname} &nbsp;·&nbsp; "
            f"Summary · {ok_count} Historical Data sheets · Methodology"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── Warnings ──────────────────────────────────────────────────────────────
    bad = [r for r in results if r["error"]]
    if bad:
        st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
        for r in bad:
            st.markdown(
                f"<div class='status-err' style='margin-bottom:6px;'>"
                f"{r['symbol']} — {r['error']}</div>",
                unsafe_allow_html=True,
            )

# ── Empty state ───────────────────────────────────────────────────────────────
elif not st.session_state.selected:
    st.markdown("""
    <div style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;
                padding:56px;text-align:center;margin-top:16px;'>
        <div style='font-size:2.4rem;margin-bottom:16px;'>📈</div>
        <div style='font-size:1.1rem;font-weight:600;color:#111827;margin-bottom:8px;'>
            Calculate equity beta for any listed company worldwide
        </div>
        <div style='font-size:0.86rem;color:#6B7280;max-width:420px;margin:0 auto;line-height:1.7;'>
            Search any company · Set analysis period · Get beta with auto-matched benchmark index
        </div>
        <div style='margin-top:32px;display:flex;justify-content:center;gap:32px;flex-wrap:wrap;'>
            <div style='text-align:center;'>
                <div style='font-size:1.4rem;font-weight:700;color:#1D4ED8;'>40+</div>
                <div style='font-size:0.7rem;color:#9CA3AF;margin-top:2px;'>Markets</div>
            </div>
            <div style='text-align:center;'>
                <div style='font-size:1.4rem;font-weight:700;color:#059669;'>Auto</div>
                <div style='font-size:0.7rem;color:#9CA3AF;margin-top:2px;'>Index Routing</div>
            </div>
            <div style='font-size:1.4rem;font-weight:700;color:#D97706;text-align:center;'>
                <div>Excel</div>
                <div style='font-size:0.7rem;color:#9CA3AF;margin-top:2px;font-weight:400;'>SLOPE Method</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
